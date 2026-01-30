from flask import Blueprint, render_template, request, jsonify, Response, send_file, url_for, send_from_directory, current_app
from datetime import datetime
import pandas as pd
import io
import zipfile
import qrcode
from io import StringIO
import csv
import os
from .manager import SimResourceManager
from models.sim_resource import SimResource, db
from .config_manager import SimConfigManager
from PIL import Image, ImageDraw, ImageFont

# 引入 OpenPyXL 樣式組件 (用於 Excel 美化)
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL_STYLES = True
except ImportError:
    HAS_OPENPYXL_STYLES = False

sim_resources_bp = Blueprint('sim_resources', __name__, url_prefix='/resources')

# SIM資源管理頁面 - 使用全寬模式
@sim_resources_bp.route('')
def resources_page():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    view_mode = request.args.get('view_mode', 'single')  # 'single' or 'range'
    default_sort = 'assigned_date' if view_mode == 'range' else 'updated_at'
    
    # 獲取搜索參數
    search_params = {
        'provider': request.args.get('provider', '').strip(),
        'card_type': request.args.get('card_type', '').strip(),
        'resources_type': request.args.get('resources_type', '').strip(),
        'batch': request.args.get('batch', '').strip(),
        'received_date': request.args.get('received_date', '').strip(),
        'imsi': request.args.get('imsi', '').strip(),
        'iccid': request.args.get('iccid', '').strip(),
        'msisdn': request.args.get('msisdn', '').strip(),
        
        # 新增參數
        'status': request.args.get('status', '').strip(),
        'customer': request.args.get('customer', '').strip(),
        'assigned_date_start': request.args.get('assigned_date_start', '').strip(),
        'assigned_date_end': request.args.get('assigned_date_end', '').strip(),
        'remark': request.args.get('remark', '').strip(),
        
        'sort': request.args.get('sort', default_sort),
        'order': request.args.get('order', 'desc'),
        'per_page': per_page
    }
    
    # 根據 view_mode 調用不同的查詢方法
    if view_mode == 'range':
        resources = SimResourceManager.get_grouped_resources(search_params, page, per_page)
    else:
        resources = SimResourceManager.get_all_resources(search_params, page, per_page)
        
    options = SimResourceManager.get_options()
    
    return render_template('resources.html', 
                          resources=resources.items, 
                          pagination=resources,
                          options=options,
                          search_params=search_params,
                          view_mode=view_mode,
                          full_width=True)

# 编辑资源路由
@sim_resources_bp.route('/api/edit/<int:resource_id>', methods=['POST'])
def edit_resource(resource_id):
    """编辑资源"""
    try:
        data = request.json
        
        # 验证数据
        errors = SimResourceManager.validate_resource_data(data, is_edit=True, resource_id=resource_id)
        if errors:
            return jsonify({'error': '; '.join(errors)}), 400
        
        # 更新资源
        resource = SimResourceManager.update_resource(resource_id, data)
        return jsonify({'success': True, 'resource': resource.to_dict()})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# 新增资源路由
@sim_resources_bp.route('/api/add', methods=['POST'])
def add_resource():
    """新增资源"""
    try:
        data = request.json
        
        # 验证数据
        errors = SimResourceManager.validate_resource_data(data)
        if errors:
            return jsonify({'error': '; '.join(errors)}), 400
        
        # 创建资源
        resource = SimResourceManager.create_resource(data)
        return jsonify({'success': True, 'resource': resource.to_dict()})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# 添加获取选项的API
@sim_resources_bp.route('/api/options')
def get_resource_options():
    """获取资源选项"""
    options = SimResourceManager.get_options()
    return jsonify(options)

# 删除资源
@sim_resources_bp.route('/api/delete/<int:resource_id>', methods=['POST'])
def delete_resource(resource_id):
    try:
        resource = SimResource.query.get_or_404(resource_id)
        db.session.delete(resource)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# 獲取導出的動態篩選選項
@sim_resources_bp.route('/api/export/options', methods=['POST'])
def get_export_filter_options():
    try:
        data = request.json
        # 從前端接收當前的搜索參數
        search_params = data.get('search_params', {})
        # 接收 Modal 內部的過濾器 (例如已選的 Customer)
        modal_filters = data.get('modal_filters', {})
        
        options = SimResourceManager.get_distinct_filters(search_params, modal_filters)
        return jsonify(options)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 导出CSV
@sim_resources_bp.route('/api/export')
def export_resources():
    try:
        resources = SimResource.query.all()
        output = []
        for res in resources:
            row = {
                'id': res.id,
                'type': res.type,
                'supplier': res.supplier,
                'created_at': res.created_at.isoformat(),
                'updated_at': res.updated_at.isoformat()
            }
            # 展开其他字段
            row.update({
                'resources_type': res.resources_type,
                'batch': res.batch,
                'received_date': res.received_date,
                'imsi': res.imsi,
                'iccid': res.iccid,
                'msisdn': res.msisdn,
                'ki': res.ki,
                'opc': res.opc,
                'lpa': res.lpa,
                'pin1': res.pin1,
                'puk1': res.puk1,
                'pin2': res.pin2,
                'puk2': res.puk2,
                'remark': res.remark
            })
            output.append(row)
        
        # 生成CSV
        si = StringIO()
        cw = csv.DictWriter(si, fieldnames=output[0].keys() if output else [])
        cw.writeheader()
        cw.writerows(output)
        return Response(si.getvalue(), mimetype='text/csv', headers={"Content-disposition": "attachment; filename=sim_resources.csv"})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# SIM Resource Import/Modify Excel
@sim_resources_bp.route('/api/import', methods=['POST'])
def import_resources():
    """
    導入資源接口 (支持 Add 新增 和 Modify 修改)
    """
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未上傳文件'}), 400
    
    file = request.files['file']
    mode = request.form.get('mode', 'add') 
    
    try:
        df = pd.read_excel(file, dtype=str)
        df = df.fillna('') 
        df.columns = df.columns.str.strip() 
        
        if mode == 'modify':
            # 但如果修改了 IMSI 本身(雖然這不常見)，也需要更新 imsi_num，不過目前模板只允許修改非主鍵
            if 'IMSI' not in df.columns:
                 return jsonify({'success': False, 'message': '模板錯誤：必須包含 "IMSI" 列 (Column A)'}), 400
            
            field_map = {
                'Ki': 'ki', 'OPC': 'opc', 'LPA': 'lpa',
                'PIN1': 'pin1', 'PUK1': 'puk1', 'PIN2': 'pin2', 'PUK2': 'puk2'
            }
            
            excel_imsis = df['IMSI'].dropna().astype(str).str.strip().tolist()
            excel_imsis = [i for i in excel_imsis if i]
            
            if not excel_imsis:
                return jsonify({'success': False, 'message': 'Excel 中沒有有效的 IMSI 數據'}), 400

            existing_resources = SimResource.query.filter(SimResource.imsi.in_(excel_imsis)).all()
            resource_map = {res.imsi: res for res in existing_resources}
            
            updated_count = 0
            not_found_count = 0
            ignored_count = 0
            
            for _, row in df.iterrows():
                imsi = str(row['IMSI']).strip()
                if not imsi: continue
                
                if imsi in resource_map:
                    resource = resource_map[imsi]
                    has_change = False
                    
                    for excel_col, db_col in field_map.items():
                        if excel_col in df.columns:
                            val = str(row[excel_col]).strip()
                            if val: 
                                setattr(resource, db_col, val)
                                has_change = True
                    
                    if has_change:
                        updated_count += 1
                    else:
                        ignored_count += 1 
                else:
                    not_found_count += 1 

            if updated_count > 0:
                db.session.commit()
                
            msg = f"修改處理完成。"
            details = []
            if updated_count > 0: details.append(f"• 成功更新: {updated_count} 筆")
            if not_found_count > 0: details.append(f"• 庫存未找到: {not_found_count} 筆 (IMSI 不存在)")
            if ignored_count > 0: details.append(f"• 未變更: {ignored_count} 筆 (未填寫修改內容)")
            
            full_msg = msg + "\n" + "\n".join(details)
                
            return jsonify({
                'success': True, 
                'message': full_msg,
                'updated_count': updated_count,
                'error_count': not_found_count
            })

        else:
            # 模式 B: 新增資源 (Add)
            required_columns = ['Provider', 'CardType', 'ResourcesType', 'Batch', 'ReceivedDate', 'IMSI', 'ICCID', 'MSISDN']
            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                return jsonify({'success': False, 'message': f'模板錯誤：缺少必要列 {", ".join(missing_cols)}'}), 400

            success_count = 0
            duplicate_count = 0
            errors = []

            imsis = df['IMSI'].dropna().astype(str).str.strip().tolist()
            iccids = df['ICCID'].dropna().astype(str).str.strip().tolist()
            
            existing_imsis = set(r[0] for r in db.session.query(SimResource.imsi).filter(SimResource.imsi.in_(imsis)).all())
            existing_iccids = set(r[0] for r in db.session.query(SimResource.iccid).filter(SimResource.iccid.in_(iccids)).all())

            new_resources = []
            for index, row in df.iterrows():
                try:
                    imsi = str(row['IMSI']).strip()
                    iccid = str(row['ICCID']).strip()
                    msisdn = str(row['MSISDN']).strip()
                    
                    if not imsi or not iccid: continue
                    
                    if imsi in existing_imsis or iccid in existing_iccids:
                        duplicate_count += 1
                        continue
                    
                    # [Optimize] 自動計算 imsi_num
                    imsi_num = int(imsi) if imsi.isdigit() else None
                    iccid_num = int(iccid) if iccid.isdigit() else None
                    msisdn_num = int(msisdn) if msisdn.isdigit() else None
                        
                    res = SimResource(
                        supplier=str(row['Provider']).strip(),
                        type=str(row['CardType']).strip(),
                        resources_type=str(row['ResourcesType']).strip(),
                        batch=str(row['Batch']).strip(),
                        received_date=str(row['ReceivedDate']).strip(),
                        imsi=imsi,
                        imsi_num=imsi_num,
                        iccid=iccid,
                        iccid_num=iccid_num,
                        msisdn=msisdn,
                        msisdn_num=msisdn_num,
                        status='Available', 
                        ki=str(row.get('Ki', '')).strip() or None,
                        opc=str(row.get('OPC', '')).strip() or None,
                        lpa=str(row.get('LPA', '')).strip() or None,
                        pin1=str(row.get('PIN1', '')).strip() or None,
                        puk1=str(row.get('PUK1', '')).strip() or None,
                        pin2=str(row.get('PIN2', '')).strip() or None,
                        puk2=str(row.get('PUK2', '')).strip() or None,
                        remark=str(row.get('Remark', '')).strip() or None
                    )
                    new_resources.append(res)
                    
                    existing_imsis.add(imsi)
                    existing_iccids.add(iccid)
                    
                except Exception as row_err:
                    errors.append(f"Row {index+2}: {str(row_err)}")

            if new_resources:
                db.session.bulk_save_objects(new_resources)
                db.session.commit()
                success_count = len(new_resources)

            msg = f"導入完成。"
            details = []
            if success_count > 0: details.append(f"• 成功新增: {success_count} 筆")
            if duplicate_count > 0: details.append(f"• 跳過重複: {duplicate_count} 筆 (IMSI/ICCID 已存在)")
            if len(errors) > 0: details.append(f"• 數據錯誤: {len(errors)} 筆")
            
            full_msg = msg + "\n" + "\n".join(details)
            
            if len(errors) > 0 and len(errors) < 5:
                full_msg += "\n\n錯誤詳情:\n" + "\n".join(errors)

            return jsonify({
                'success': True,
                'message': full_msg,
                'success_count': success_count,
                'duplicate_count': duplicate_count,
                'error_count': len(errors)
            })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f"文件處理失敗: {str(e)}"}), 500

# 用於下載修改模板的路由
@sim_resources_bp.route('/api/template/<filename>')
def download_template(filename):
    """下載 Excel 模板"""
    # 假設您的模板存放在 app 根目錄下的 ExcelTemplate 文件夾
    # 您可能需要根據實際目錄結構調整 directory 參數
    template_dir = os.path.join(current_app.root_path, 'ExcelTemplate')
    
    # 安全檢查，只允許下載特定的模板
    allowed_templates = ['SIM_Resource_Template.xlsx', 'SIM_Resource_Modify_Template.xlsx']
    if filename not in allowed_templates:
        return jsonify({'success': False, 'message': '文件不存在或不允許下載'}), 404
        
    try:
        return send_from_directory(directory=template_dir, path=filename, as_attachment=True)
    except FileNotFoundError:
        return jsonify({'success': False, 'message': '服務器上找不到模板文件'}), 404

@sim_resources_bp.route('/api/assign/calculate', methods=['POST'])
def calculate_assignment():
    """計算分配選項"""
    data = request.json
    result = SimResourceManager.calculate_assignment_options(
        provider=data.get('provider'),
        card_type=data.get('card_type'),
        resources_type=data.get('resources_type'),
        quantity=int(data.get('quantity', 0))
    )
    return jsonify(result)

@sim_resources_bp.route('/api/assign/confirm', methods=['POST'])
def confirm_assignment():
    """確認並執行自動分配"""
    data = request.json
    result = SimResourceManager.confirm_assignment(
        plan=data.get('batches'),
        customer=data.get('customer'),
        assigned_date=data.get('assigned_date'),
        remark=data.get('remark'),
        provider=data.get('provider'),
        card_type=data.get('card_type'),
        resources_type=data.get('resources_type')
    )
    return jsonify(result)

@sim_resources_bp.route('/api/assign/manual', methods=['POST'])
def manual_assignment():
    """執行手動分配 (支持範圍或已選)"""
    data = request.json
    result = SimResourceManager.manual_assignment(
        scope=data.get('scope'),
        ids=data.get('ids'),
        start_imsi=data.get('start_imsi'),
        end_imsi=data.get('end_imsi'),
        customer=data.get('customer'),
        assigned_date=data.get('assigned_date'),
        remark=data.get('remark')
    )
    return jsonify(result)

@sim_resources_bp.route('/api/assign/cancel', methods=['POST'])
def cancel_assignment():
    """執行批量取消分配 (支持範圍或已選)"""
    data = request.json
    # 注意：這裡改用新的 batch_cancel_assignment 方法
    result = SimResourceManager.batch_cancel_assignment(
        scope=data.get('scope'),
        ids=data.get('ids'),
        start_imsi=data.get('start_imsi'),
        end_imsi=data.get('end_imsi'),
        remark=data.get('remark')
    )
    return jsonify(result)

# 導出接口
@sim_resources_bp.route('/api/export_custom', methods=['POST'])
def export_custom_resources():
    """
    自定義導出接口 (最終美化版：Excel 格式美化 + QR Zip 分離下載)
    """
    try:
        data = request.json
        
        scope = data.get('scope')
        selected_ids = data.get('selected_ids')
        search_params = data.get('search_params')
        selected_columns = data.get('columns', [])
        filter_customer = data.get('filter_customer')
        filter_assigned_date = data.get('filter_assigned_date')
        
        include_qrcode = data.get('include_qrcode', False)
        only_qrcode = data.get('only_qrcode', False)
        
        # 傳遞給 Manager (這裡已經在 manager.py 中修復了對 Range Object 的處理)
        resources = SimResourceManager.get_resources_for_export(
            scope=scope,
            selected_ids=selected_ids,
            search_params=search_params,
            extra_filters={
                'customer': filter_customer,
                'assigned_date': filter_assigned_date
            }
        )
        
        # 全局數量限制
        if len(resources) > 10000:
            return jsonify({'error': f'導出數量限制為 10,000 筆。當前篩選結果共 {len(resources)} 筆，請縮小範圍。'}), 400
        
        if not resources:
            return jsonify({'error': '沒有符合條件的數據可導出'}), 400

        # =========================================================
        # 分支 A: 導出 Excel (美化版)
        # =========================================================
        if not only_qrcode:
            # 檢查是否具備美化所需的庫
            if not HAS_OPENPYXL_STYLES:
                return jsonify({'error': "導出失敗: 缺少 'openpyxl' 庫，無法執行格式美化。請聯繫管理員執行 'pip install openpyxl'"}), 500

            export_list = []
            for res in resources:
                row = {
                    'IMSI': res.imsi, 'ICCID': res.iccid, 'MSISDN': res.msisdn, 'LPA': res.lpa,
                    'Ki': res.ki, 'OPC': res.opc, 'Customer': res.customer, 'Assign Date': res.assigned_date,
                    'Provider': res.supplier, 'CardType': res.type, 'ResourcesType': res.resources_type,
                    'Remark': res.remark, 'Status': res.status, 'Batch': res.batch, 'ReceivedDate': res.received_date,
                    'PIN1': res.pin1, 'PUK1': res.puk1, 'PIN2': res.pin2, 'PUK2': res.puk2,
                    'Created At': res.created_at, 'Updated At': res.updated_at
                }
                filtered_row = {k: v for k, v in row.items() if k in selected_columns}
                export_list.append(filtered_row)

            df = pd.DataFrame(export_list)
            excel_io = io.BytesIO()
            
            # 使用 openpyxl 引擎進行寫入
            with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Resources')
                
                # === 開始美化 Excel ===
                workbook = writer.book
                worksheet = writer.sheets['Resources']
                
                # 1. 定義樣式對象
                # 字體: Calibri, Size 10
                font_style = Font(name='Calibri', size=10)
                # 標題字體: Calibri, Size 10, 粗體
                font_header = Font(name='Calibri', size=10, bold=True)
                
                # 邊框: 四周細線
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                
                # 對齊: 靠左+垂直置中
                align_data = Alignment(horizontal='left', vertical='center', wrap_text=False)
                align_header = Alignment(horizontal='left', vertical='center', wrap_text=False)
                
                # 填充: 標題黃色 (Solid Yellow)
                fill_header = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # 2. 遍歷所有單元格應用樣式
                max_row = worksheet.max_row
                max_col = worksheet.max_column

                # 用於計算列寬
                col_widths = {}

                # iter_rows 是一個生成器，逐行處理
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        cell.border = thin_border  # 全局加上邊框
                        
                        # 獲取內容長度用於計算列寬
                        val = str(cell.value) if cell.value else ""
                        # 計算字元長度 (稍微加寬一點 buffer)
                        # 注意：中文字符可能需要額外計算，這裡做簡單長度估計
                        length = len(val)
                        # 如果是標題行，長度要考慮過濾箭頭的空間，多加一點
                        if row_idx == 1:
                            length += 4 
                            
                        current_w = col_widths.get(col_idx, 0)
                        col_widths[col_idx] = max(current_w, length)

                        # 區分標題行和內容行
                        if row_idx == 1:
                            # 標題行樣式
                            cell.font = font_header
                            cell.alignment = align_header
                            cell.fill = fill_header
                        else:
                            # 內容行樣式
                            cell.font = font_style
                            cell.alignment = align_data
                
                # 3. 設置自動列寬
                for col_idx, width in col_widths.items():
                    col_letter = get_column_letter(col_idx)
                    # 限制最大寬度為 60，最小寬度為 10 (避免太窄或太寬)
                    # 基礎係數 1.2 讓顯示更舒適
                    adjusted_width = min(max(width * 1.2, 10), 60)
                    worksheet.column_dimensions[col_letter].width = adjusted_width

            excel_io.seek(0)
            return send_file(
                excel_io,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='sim_resources.xlsx'
            )

        # =========================================================
        # 分支 B: 導出 ZIP (僅包含 QR Codes)
        # =========================================================
        else:
            zip_io = io.BytesIO()
            
            # 使用標準 zipfile
            with zipfile.ZipFile(zip_io, 'w', zipfile.ZIP_DEFLATED) as zf:
                qr_factory = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=2)
                
                # 設定字體大小
                target_font_size = 26
                font = None
                
                # 嘗試加載系統字體
                font_candidates = [
                    "arial.ttf",  # Windows
                    "Arial.ttf",
                    "DejaVuSans.ttf", # Linux
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                    "/usr/share/fonts/TTF/DejaVuSans.ttf"
                ]
                
                for font_path in font_candidates:
                    try:
                        font = ImageFont.truetype(font_path, target_font_size)
                        break
                    except (OSError, IOError):
                        continue
                
                if font is None:
                    print("Warning: No system font found, falling back to default.")
                    font = ImageFont.load_default()

                for res in resources:
                    if res.type == 'eSIM' and res.lpa:
                        try:
                            qr_factory.clear()
                            qr_factory.add_data(res.lpa)
                            qr_factory.make(fit=True)
                            img = qr_factory.make_image(fill_color="black", back_color="white").get_image()
                            
                            # === 添加 ICCID 文字 (置中優化版) ===
                            if res.iccid:
                                txt = res.iccid
                                draw = ImageDraw.Draw(img)
                                
                                # 精確計算文字邊界
                                if hasattr(draw, 'textbbox'):
                                    bbox = draw.textbbox((0, 0), txt, font=font)
                                    text_width = bbox[2] - bbox[0]
                                    text_height = bbox[3] - bbox[1]
                                    # 修正：加上 descent (基線以下的高度)，讓視覺更平衡
                                    text_height += bbox[3] 
                                else:
                                    text_width, text_height = draw.textsize(txt, font=font)
                                
                                # 設定上下間距，確保文字在白條中間
                                padding_top = 1
                                padding_bottom = 1
                                
                                # 計算底部白條的總高度
                                footer_height = text_height + padding_top + padding_bottom
                                
                                # 創建新畫布 (總高度 = 原圖高 + 底部白條高)
                                new_height = img.height + footer_height
                                new_img = Image.new('RGB', (img.width, new_height), 'white')
                                
                                # 1. 貼上原始 QR Code
                                new_img.paste(img, (0, 0))
                                
                                # 2. 繪製文字 (水平置中 + 垂直置中)
                                draw_new = ImageDraw.Draw(new_img)
                                
                                # 水平位置 (置中)
                                text_x = (img.width - text_width) // 2
                                text_x = max(0, text_x)
                                
                                # 垂直位置 (原圖高度 + 上方間距)
                                # 這樣文字就會位於底部白條的正中間
                                text_y = img.height + padding_top
                                
                                draw_new.text((text_x, text_y), txt, font=font, fill="black")
                                
                                img = new_img
                            # ============================
                            
                            img_byte_arr = io.BytesIO()
                            img.save(img_byte_arr, format='PNG')
                            
                            if res.iccid: fname = f"{res.iccid}.png"
                            elif res.imsi: fname = f"{res.imsi}.png"
                            else: fname = f"unknown_{res.id}.png"
                            
                            zf.writestr(fname, img_byte_arr.getvalue())
                            
                        except Exception as e:
                            print(f"QR Gen Error Resource ID {res.id}: {e}")
                            continue
            
            zip_io.seek(0)
            return send_file(
                zip_io,
                mimetype='application/zip',
                as_attachment=True,
                download_name='qrcodes_package.zip'
            )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'導出處理失敗: {str(e)}'}), 500
    
@sim_resources_bp.route('/api/batch/operation', methods=['POST'])
def batch_operation():
    """執行批量操作 (編輯/刪除)"""
    try:
        data = request.json
        action = data.get('action') # 'edit' 或 'delete'
        scope = data.get('scope')   # 'selected' 或 'range'
        ids = data.get('ids', [])
        start_imsi = data.get('start_imsi')
        end_imsi = data.get('end_imsi')
        
        if action == 'edit':
            update_data = data.get('data', {})
            result = SimResourceManager.batch_update_resources(scope, ids, start_imsi, end_imsi, update_data)
        elif action == 'delete':
            result = SimResourceManager.batch_delete_resources(scope, ids, start_imsi, end_imsi)
        else:
            return jsonify({"success": False, "message": "無效的操作類型"}), 400
            
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@sim_resources_bp.route('/api/inventory_stats')
def get_inventory_stats():
    """獲取庫存統計 API"""
    stats = SimResourceManager.get_inventory_stats()
    
    # 動態獲取閾值
    config = SimConfigManager.load_config()
    threshold = config.get('low_stock_threshold', 1000)
    
    return jsonify({
        'stats': stats,
        'threshold': threshold
    })  
    
@sim_resources_bp.route('/api/config/get', methods=['GET'])
def get_config():
    """獲取當前配置"""
    try:
        config = SimConfigManager.load_config()
        return jsonify(config)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@sim_resources_bp.route('/api/config/save', methods=['POST'])
def save_config():
    """保存配置"""
    try:
        new_config = request.json
        SimConfigManager.save_config(new_config)
        return jsonify({'success': True, 'message': '配置已保存'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@sim_resources_bp.route('/api/config/check_usage', methods=['POST'])
def check_config_usage():
    """檢查配置項是否被使用"""
    try:
        data = request.json
        category = data.get('category')
        value = data.get('value')
        
        is_used = SimConfigManager.check_usage(category, value)
        return jsonify({'used': is_used})
    except Exception as e:
        return jsonify({'used': True, 'error': str(e)}), 500 # 出錯時默認當作被使用，防止誤刪    
    
# 批量操作: 按導入IMSI
@sim_resources_bp.route('/api/batch/resolve_imsis', methods=['POST'])
def resolve_imsis_for_batch():
    """解析批量操作導入的 IMSI Excel (性能優化版)"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未上傳文件'}), 400
    
    file = request.files['file']
    try:
        # 1. 讀取 Excel (只讀取需要的數據)
        # header=None 先讀取所有內容，稍後判斷
        df = pd.read_excel(file, header=None, dtype=str)
        
        target_series = pd.Series(dtype=str)
        
        # 智能尋找數據列
        # 嘗試 A: 尋找包含 "IMSI" 字樣的行作為標題
        # 展平所有數據並轉為字符串，去除空格
        all_values = df.values.flatten().astype(str)
        all_values = [x.strip() for x in all_values if x.lower() != 'nan' and x.lower() != 'imsi']
        
        # 使用 Pandas 向量化操作進行過濾 (比 Python for loop 快很多)
        s_values = pd.Series(all_values)
        
        # 2. 格式驗證 (15位數字)
        # 使用正則表達式匹配 15 位數字
        valid_mask = s_values.str.match(r'^\d{15}$')
        
        # 提取有效 IMSI (去重)
        valid_imsis = s_values[valid_mask].unique().tolist()
        
        # 統計無效數據 (這裡簡單計算總數差，不一一列出以節省資源)
        total_count = len(s_values)
        valid_count = len(valid_imsis)
        invalid_count = total_count - valid_count
        
        if not valid_imsis:
             return jsonify({'success': False, 'message': f'未在文件中找到有效的 15 位 IMSI 號碼 (共掃描 {total_count} 個單元格)。'}), 400

        # 3. 數據庫查詢優化
        # 只查詢 id 和 imsi 欄位，不加載整個對象
        # 使用 yield_per 分批處理，防止內存溢出 (雖然這裡只查 ID 影響不大，但好習慣)
        found_records = db.session.query(SimResource.imsi, SimResource.id)\
            .filter(SimResource.imsi.in_(valid_imsis))\
            .all()
        
        # 4. 構建結果
        found_map = {rec.imsi: rec.id for rec in found_records}
        found_ids = list(found_map.values())
        
        # 計算未找到的 IMSI
        not_found_count = len(valid_imsis) - len(found_ids)
        
        # 構建訊息
        message = f"解析完成！"
        details = [f"• 成功匹配庫存: {len(found_ids)} 筆"]
        
        if invalid_count > 0:
            details.append(f"• 忽略格式不符: {invalid_count} 筆 (非15位數字或標題)")
        
        if not_found_count > 0:
            details.append(f"• 庫存未找到: {not_found_count} 筆")
            
        full_message = message + "\n" + "\n".join(details)
            
        return jsonify({
            'success': True,
            'ids': found_ids,
            'count': len(found_ids),
            'message': full_message,
            'invalid_count': invalid_count,
            'not_found_count': not_found_count
        })

    except Exception as e:
        # import traceback
        # traceback.print_exc()
        return jsonify({'success': False, 'message': f"解析失敗: {str(e)}"}), 500
    
# 用於即時查看單個資源 QR Code 的接口
@sim_resources_bp.route('/api/qrcode/view/<int:resource_id>')
def get_resource_qrcode(resource_id):
    """即時生成並返回單個資源的 QR Code 圖片流"""
    resource = SimResource.query.get_or_404(resource_id)
    
    if not resource.lpa:
        return jsonify({'success': False, 'message': '此資源沒有 LPA 數據'}), 404
        
    try:
        # 生成 QR Code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L, # 使用低容錯率 (L) 以減小體積和加快生成速度
            box_size=10,
            border=2,
        )
        qr.add_data(resource.lpa)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        # 轉為字節流
        img_io = io.BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)
        
        return send_file(img_io, mimetype='image/png')
    except Exception as e:
        return jsonify({'success': False, 'message': f'生成失敗: {str(e)}'}), 500    